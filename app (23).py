import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
import base64, textwrap, tempfile, os

st.set_page_config(page_title="Collections Hourly Snapshot", layout="wide")
st.title("Collections Hourly Snapshot Dashboard")

# ── TIME SLOTS / DAY-PROCESS CONFIG ─────────────────────────────────────────────
TIME_SLOTS = ["9 AM", "12 PM", "3 PM", "6 PM", "9 PM", "12 AM", "3 AM", "6 AM"]
DAY_PROCESS_BUCKETS = ["FRESH", "1-29", "30-59", "60-89"]
DAY_PROCESS_BUCKET_KEY = {"FRESH": "Fresh", "1-29": "1-29", "30-59": "30-59", "60-89": "60-89"}

uploaded_file = st.file_uploader("1️⃣ Upload Raw Source Excel (.xlsx)", type=["xlsx"], key="raw_file")
prev_file = st.file_uploader(
    "2️⃣ Upload Previously Downloaded Snapshot (optional — carries forward earlier time-slot data)",
    type=["xlsx"], key="prev_file",
)


# ── PARSE ──────────────────────────────────────────────────────────────────────
def parse_sheet3(wb):
    if len(wb.worksheets) < 3:
        raise ValueError(
            "This workbook doesn't have a 3rd sheet with bucket/zone data. "
            "Please make sure you uploaded the original raw source Excel in field 1, "
            "not a previously downloaded snapshot."
        )
    ws  = wb.worksheets[2]
    rows = list(ws.iter_rows(values_only=True))
    eff_row, flow_row = rows[2], rows[3]

    bucket_cols = {
        'Fresh':             2,
        '1-29':              4,
        '1-29 NORM':         6,
        '30-59':             8,
        '60-89':             10,
        'STAGE 2 Reduction': 12,
    }
    bucket_data = {}
    for name, c in bucket_cols.items():
        bucket_data[name] = {
            'eff':  eff_row[c]  or 0,
            'flow': flow_row[c] or 0,
        }

    zone_data = []
    for row in rows[8:23]:
        if row[1] is None:
            continue
        zone_data.append({
            'zone':  row[1],
            'fresh': row[2],
            '1-29':  row[3],
            '30-59': row[4],
            '60-89': row[5],
        })
    return bucket_data, zone_data

def get_val(zone_data, name, bucket):
    for r in zone_data:
        if r['zone'].strip().upper() == name.strip().upper():
            v = r.get(bucket)
            return f"{v*100:.2f}%" if v is not None else "—"
    return "—"

def get_raw_val(zone_data, name, bucket):
    for r in zone_data:
        if r['zone'].strip().upper() == name.strip().upper():
            return r.get(bucket)
    return None

def pct(v):
    return f"{v*100:.2f}%" if v else "—"

def flow_fmt(v):
    return f"₹ {v:,.2f} Crs" if v else "—"


# ── DAY-PROCESS HISTORY (the persistence mechanism) ─────────────────────────────
def load_day_process_history(prev_excel_file):
    """Reads the 'DayProcess' sheet of a previously downloaded snapshot (if any)
    and returns the carried-forward per-time-slot values."""
    history = {b: {t: None for t in TIME_SLOTS} for b in DAY_PROCESS_BUCKETS}
    if prev_excel_file is None:
        return history
    try:
        prev_wb = openpyxl.load_workbook(prev_excel_file, data_only=True)
    except Exception:
        return history
    if "DayProcess" not in prev_wb.sheetnames:
        return history
    ws = prev_wb["DayProcess"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return history
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_for_time = {t: header.index(t) for t in TIME_SLOTS if t in header}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        bucket_name = str(row[0]).strip().upper()
        if bucket_name not in DAY_PROCESS_BUCKETS:
            continue
        for t, idx in col_for_time.items():
            if idx < len(row):
                val = row[idx]
                if isinstance(val, (int, float)):
                    history[bucket_name][t] = float(val)
    return history

def update_day_process_history(history, bucket_data, selected_time):
    """Returns a new history dict with the selected time-slot filled in from
    the freshly-parsed raw source file, leaving every other slot untouched."""
    updated = {b: dict(v) for b, v in history.items()}
    for disp_name, key in DAY_PROCESS_BUCKET_KEY.items():
        eff = bucket_data.get(key, {}).get('eff', 0) or 0
        updated[disp_name][selected_time] = eff
    return updated

def day_process_total(history, t):
    vals = [history[b][t] for b in DAY_PROCESS_BUCKETS if history[b].get(t) is not None]
    if not vals:
        return None
    return sum(vals) / len(vals)


# ── HTML BUILDER ───────────────────────────────────────────────────────────────
ZONE_ORDER = [
    ('PAN INDIA', []),
    ('NORTH',     ['NORTH_1','NORTH_2','NORTH_3']),
    ('SOUTH',     ['SOUTH_1','SOUTH_2']),
    ('WEST',      ['WEST_1','WEST_2']),
    ('EAST',      ['EAST_1','EAST_2']),
]

BOX_CONFIGS = [
    ('Fresh',             'Fresh',             '#1a6080', '#ffffff'),
    ('1-29',              '1-29 (S2 Concern)', '#e6a817', '#ffffff'),
    ('1-29 NORM',         '1-29 Norm',         '#e6a817', '#ffffff'),
    ('30-59',             '30-59',             '#3da850', '#ffffff'),
    ('60-89',             '60-89 (S3 Concern)','#c8846e', '#ffffff'),
    ('STAGE 2 Reduction', 'S2 Roll Back',      '#c8846e', '#ffffff'),
]

def build_day_process_table_html(history, times):
    header_cells = "".join(f"<th>{t}</th>" for t in times)
    rows_html = ""
    for i, b in enumerate(DAY_PROCESS_BUCKETS):
        alt = "dp-row-alt" if i % 2 == 0 else ""
        cells = "".join(f"<td>{pct(history[b].get(t))}</td>" for t in times)
        rows_html += f"""
        <tr class="{alt}">
          <td class="dp-bucket"><b>{b}</b></td>
          {cells}
        </tr>"""
    total_cells = "".join(f"<td>{pct(day_process_total(history, t))}</td>" for t in times)
    rows_html += f"""
        <tr class="dp-total">
          <td class="dp-bucket"><b>TOTAL</b></td>
          {total_cells}
        </tr>"""
    return f"""
    <table class="dp-table">
      <thead><tr><th>BUCKET</th>{header_cells}</tr></thead>
      <tbody>{rows_html}</tbody>
    </table>"""

def build_html(bucket_data, zone_data, timestamp_str, history):
    # ── top boxes
    boxes_html = ""
    for key, label, bg, tc in BOX_CONFIGS:
        bd   = bucket_data.get(key, {})
        eff  = bd.get('eff', 0)
        fl   = bd.get('flow', 0)
        boxes_html += f"""
        <div class="box">
          <div class="box-header" style="background:{bg};color:{tc};">{label}</div>
          <div class="box-body">
            <div class="box-eff">{pct(eff)}</div>
            <div class="box-lmtd">vs. LMTD &nbsp;</div>
            <div class="box-flow">Flow Value &nbsp;<b>{fl:,.2f} Crs</b></div>
          </div>
        </div>"""

    # ── sub-zone rows
    zone_rows_html = ""
    for zone_name, subs in ZONE_ORDER:
        display = zone_name.replace('_',' ')
        bg_class = "row-pan" if zone_name == "PAN INDIA" else "row-zone"
        zone_rows_html += f"""
        <tr class="{bg_class}">
          <td class="zone-name"><b>{display}</b></td>
          <td>{get_val(zone_data, zone_name, 'fresh')}</td>
          <td>{get_val(zone_data, zone_name, '1-29')}</td>
          <td>{get_val(zone_data, zone_name, '30-59')}</td>
          <td>{get_val(zone_data, zone_name, '60-89')}</td>
        </tr>"""
        for i, sub in enumerate(subs):
            disp_sub = sub.replace('_',' ')
            alt = "row-alt" if i % 2 == 1 else ""
            zone_rows_html += f"""
        <tr class="row-sub {alt}">
          <td class="sub-name">{disp_sub}</td>
          <td>{get_val(zone_data, sub, 'fresh')}</td>
          <td>{get_val(zone_data, sub, '1-29')}</td>
          <td>{get_val(zone_data, sub, '30-59')}</td>
          <td>{get_val(zone_data, sub, '60-89')}</td>
        </tr>"""

    # ── day process tables (two stacked: 9AM-6PM and 9PM-6AM)
    dp_table_1 = build_day_process_table_html(history, TIME_SLOTS[:4])
    dp_table_2 = build_day_process_table_html(history, TIME_SLOTS[4:])

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8"/>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Segoe UI', Arial, sans-serif;
    background: #f0f2f5;
    padding: 24px;
    width: 1400px;
  }}
  .card {{
    background: white;
    border-radius: 10px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.10);
    padding: 22px 26px 26px 26px;
  }}
  .title {{
    font-size: 20px;
    font-weight: 700;
    color: #1a1a2e;
    margin-bottom: 20px;
    letter-spacing: 0.3px;
  }}
  /* ── TOP BOXES */
  .boxes {{
    display: flex;
    gap: 14px;
    margin-bottom: 24px;
  }}
  .box {{
    flex: 1;
    border-radius: 8px;
    border: 1px solid #dde3ea;
    overflow: hidden;
    box-shadow: 0 1px 4px rgba(0,0,0,0.07);
  }}
  .box-header {{
    padding: 9px 12px;
    font-size: 13px;
    font-weight: 700;
    text-align: center;
    letter-spacing: 0.2px;
  }}
  .box-body {{
    padding: 12px 10px 14px 10px;
    text-align: center;
    background: white;
  }}
  .box-eff {{
    font-size: 22px;
    font-weight: 800;
    color: #1a1a2e;
    line-height: 1.2;
  }}
  .box-lmtd {{
    font-size: 11px;
    color: #888;
    margin: 5px 0 6px 0;
  }}
  .box-flow {{
    font-size: 11.5px;
    color: #444;
  }}
  /* ── TABLE (shared) */
  .section-title {{
    font-size: 13px;
    font-weight: 700;
    color: #1a3a5c;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    margin-bottom: 10px;
    padding-bottom: 6px;
    border-bottom: 2px solid #1a6080;
    display: inline-block;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 12.5px;
  }}
  thead th {{
    background: #1a6080;
    color: white;
    padding: 9px 12px;
    text-align: center;
    font-weight: 700;
    font-size: 12px;
    letter-spacing: 0.4px;
  }}
  thead th:first-child {{ text-align: left; padding-left: 16px; }}
  td {{
    padding: 8px 12px;
    text-align: center;
    color: #333;
    border-bottom: 1px solid #eef0f3;
  }}
  td:first-child {{ text-align: left; }}
  .row-pan   {{ background: #e6f4ea; }}
  .row-pan td {{ color: #1a5c2a; font-size: 13px; }}
  .row-zone  {{ background: #dce8f0; }}
  .row-zone td {{ color: #1a3a5c; }}
  .zone-name {{ padding-left: 16px !important; }}
  .row-sub td {{ color: #444; }}
  .row-alt   {{ background: #f5f8fb; }}
  .sub-name  {{ padding-left: 30px !important; color: #555 !important; }}
  tr:hover   {{ background: #eef5fb !important; transition: background 0.15s; }}

  /* ── BOTTOM TWO-COLUMN LAYOUT */
  .bottom-row {{
    display: flex;
    gap: 24px;
    align-items: flex-start;
  }}
  .col-subzone {{ flex: 1.5; min-width: 0; }}
  .col-dayprocess {{ flex: 1; min-width: 0; }}

  /* ── DAY PROCESS MINI TABLES */
  .dp-table {{
    font-size: 11.5px;
    margin-top: 6px;
  }}
  .dp-table th {{
    padding: 6px 8px;
    font-size: 10.5px;
  }}
  .dp-table th:first-child {{ padding-left: 10px; }}
  .dp-table td {{ padding: 6px 8px; }}
  .dp-table td:first-child {{ padding-left: 10px !important; }}
  .dp-bucket {{ text-align: left; color: #1a3a5c; }}
  .dp-row-alt {{ background: #dce8f0; }}
  .dp-total td {{
    font-weight: 700;
    color: #1a1a2e;
    border-top: 2px solid #1a3a5c;
    border-bottom: none;
  }}
</style>
</head>
<body>
<div class="card">
  <div class="title">Collections Hourly Snapshot &nbsp;|&nbsp; Updates &nbsp;|&nbsp; {timestamp_str}</div>

  <div class="boxes">{boxes_html}</div>

  <div class="bottom-row">
    <div class="col-subzone">
      <div class="section-title">Sub-Zone Cut (Efficiency%)</div>
      <table>
        <thead>
          <tr>
            <th style="width:22%">ZONE</th>
            <th>FRESH</th>
            <th>1-29</th>
            <th>30-59</th>
            <th>60-89</th>
          </tr>
        </thead>
        <tbody>
          {zone_rows_html}
        </tbody>
      </table>
    </div>

    <div class="col-dayprocess">
      <div class="section-title">Day Process – Efficiency%</div>
      {dp_table_1}
      {dp_table_2}
    </div>
  </div>
</div>
</body>
</html>"""
    return html


def render_dashboard_png(bucket_data, zone_data, timestamp_str, history) -> bytes:
    """Renders the full dashboard (boxes + sub-zone cut + day process) straight to a
    PNG using Matplotlib — no browser/Chromium needed."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.patches import Rectangle

    W = 1450
    PAD = 24
    CARD_PAD = 26
    inner_w = W - 2 * PAD - 2 * CARD_PAD

    box_gap   = 14
    n_boxes   = len(BOX_CONFIGS)
    box_w     = (inner_w - box_gap * (n_boxes - 1)) / n_boxes
    box_h     = 95
    box_hdr_h = 30

    title_h           = 30
    gap_after_title    = 14
    gap_after_boxes    = 24
    section_h          = 30
    gap_after_section  = 10
    col_gap            = 24

    left_w  = (inner_w - col_gap) * 0.58
    right_w = (inner_w - col_gap) * 0.42

    # left column (sub-zone cut) height
    sz_total_rows = sum(1 + len(subs) for _, subs in ZONE_ORDER)
    sz_row_h    = 30
    sz_header_h = 34
    left_col_h  = section_h + gap_after_section + sz_header_h + sz_total_rows * sz_row_h

    # right column (day process) height — two stacked mini tables
    dp_row_h       = 26
    dp_header_h    = 26
    dp_table_h     = dp_header_h + (len(DAY_PROCESS_BUCKETS) + 1) * dp_row_h
    dp_gap_between = 14
    right_col_h    = section_h + gap_after_section + dp_table_h * 2 + dp_gap_between

    table_area_h = max(left_col_h, right_col_h)

    H = (PAD * 2 + CARD_PAD * 2 + title_h + gap_after_title + box_h + gap_after_boxes
         + table_area_h + 10)

    fig, ax = plt.subplots(figsize=(W / 100, H / 100), dpi=200)
    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)
    ax.set_xlim(0, W)
    ax.set_ylim(0, H)
    ax.invert_yaxis()
    ax.axis("off")
    fig.patch.set_facecolor("#f0f2f5")

    ax.add_patch(Rectangle((PAD, PAD), W - 2 * PAD, H - 2 * PAD,
                            facecolor="white", edgecolor="none", zorder=1))

    x0 = PAD + CARD_PAD
    y  = PAD + CARD_PAD

    ax.text(x0, y, f"Collections Hourly Snapshot  |  Updates  |  {timestamp_str}",
            fontsize=14, fontweight="bold", color="#1a1a2e", ha="left", va="top", zorder=3)
    y += title_h + gap_after_title

    # top boxes
    bx = x0
    for key, label, bg, tc in BOX_CONFIGS:
        bd  = bucket_data.get(key, {})
        eff = bd.get('eff', 0)
        fl  = bd.get('flow', 0)
        ax.add_patch(Rectangle((bx, y), box_w, box_h, facecolor="white",
                                edgecolor="#dde3ea", linewidth=1, zorder=2))
        ax.add_patch(Rectangle((bx, y), box_w, box_hdr_h, facecolor=bg,
                                edgecolor="none", zorder=3))
        ax.text(bx + box_w / 2, y + box_hdr_h / 2, label, fontsize=9.5, fontweight="bold",
                color=tc, ha="center", va="center", zorder=4)
        ax.text(bx + box_w / 2, y + box_hdr_h + 24, pct(eff), fontsize=16, fontweight="bold",
                color="#1a1a2e", ha="center", va="center", zorder=4)
        ax.text(bx + box_w / 2, y + box_hdr_h + 44, "vs. LMTD", fontsize=8,
                color="#888888", ha="center", va="center", zorder=4)
        ax.text(bx + box_w / 2, y + box_hdr_h + 62, f"Flow Value  {fl:,.2f} Crs", fontsize=8,
                color="#444444", ha="center", va="center", zorder=4)
        bx += box_w + box_gap

    y += box_h + gap_after_boxes
    table_top = y

    # ---- LEFT: sub-zone cut ----
    lx = x0
    ax.text(lx, table_top, "SUB-ZONE CUT (EFFICIENCY%)", fontsize=9.5, fontweight="bold",
            color="#1a3a5c", ha="left", va="top", zorder=3)
    ax.plot([lx, lx + 230], [table_top + section_h - 8, table_top + section_h - 8],
            color="#1a6080", linewidth=2, zorder=3)
    ly = table_top + section_h + gap_after_section

    col_w0     = left_w * 0.26
    col_w_rest = (left_w - col_w0) / 4
    col_x = [lx, lx + col_w0, lx + col_w0 + col_w_rest,
             lx + col_w0 + col_w_rest * 2, lx + col_w0 + col_w_rest * 3]
    headers = ["ZONE", "FRESH", "1-29", "30-59", "60-89"]

    ax.add_patch(Rectangle((lx, ly), left_w, sz_header_h, facecolor="#1a6080",
                            edgecolor="none", zorder=2))
    ax.text(lx + 12, ly + sz_header_h / 2, headers[0], fontsize=8.5, fontweight="bold",
            color="white", ha="left", va="center", zorder=3)
    for i in range(1, 5):
        cx = col_x[i] + col_w_rest / 2
        ax.text(cx, ly + sz_header_h / 2, headers[i], fontsize=8.5, fontweight="bold",
                color="white", ha="center", va="center", zorder=3)
    ly += sz_header_h

    for zone_name, subs in ZONE_ORDER:
        display  = zone_name.replace('_', ' ')
        is_pan   = zone_name == "PAN INDIA"
        row_bg   = "#e6f4ea" if is_pan else "#dce8f0"
        tcolor   = "#1a5c2a" if is_pan else "#1a3a5c"
        ax.add_patch(Rectangle((lx, ly), left_w, sz_row_h, facecolor=row_bg,
                                edgecolor="none", zorder=2))
        ax.text(lx + 12, ly + sz_row_h / 2, display, fontsize=8.5, fontweight="bold",
                color=tcolor, ha="left", va="center", zorder=3)
        vals = [get_val(zone_data, zone_name, b) for b in ('fresh', '1-29', '30-59', '60-89')]
        for i, v in enumerate(vals):
            cx = col_x[i + 1] + col_w_rest / 2
            ax.text(cx, ly + sz_row_h / 2, v, fontsize=8.5, color=tcolor,
                    ha="center", va="center", zorder=3)
        ax.plot([lx, lx + left_w], [ly + sz_row_h, ly + sz_row_h], color="#eef0f3", linewidth=1, zorder=3)
        ly += sz_row_h

        for i, sub in enumerate(subs):
            disp_sub = sub.replace('_', ' ')
            alt      = (i % 2 == 1)
            row_bg   = "#f5f8fb" if alt else "#ffffff"
            ax.add_patch(Rectangle((lx, ly), left_w, sz_row_h, facecolor=row_bg,
                                    edgecolor="none", zorder=2))
            ax.text(lx + 24, ly + sz_row_h / 2, disp_sub, fontsize=8.5, color="#555555",
                    ha="left", va="center", zorder=3)
            vals = [get_val(zone_data, sub, b) for b in ('fresh', '1-29', '30-59', '60-89')]
            for j, v in enumerate(vals):
                cx = col_x[j + 1] + col_w_rest / 2
                ax.text(cx, ly + sz_row_h / 2, v, fontsize=8.5, color="#444444",
                        ha="center", va="center", zorder=3)
            ax.plot([lx, lx + left_w], [ly + sz_row_h, ly + sz_row_h], color="#eef0f3", linewidth=1, zorder=3)
            ly += sz_row_h

    # ---- RIGHT: day process ----
    rx = x0 + left_w + col_gap
    ax.text(rx, table_top, "DAY PROCESS – EFFICIENCY%", fontsize=9.5, fontweight="bold",
            color="#1a3a5c", ha="left", va="top", zorder=3)
    ax.plot([rx, rx + 230], [table_top + section_h - 8, table_top + section_h - 8],
            color="#1a6080", linewidth=2, zorder=3)
    ry = table_top + section_h + gap_after_section

    dp_bucket_w = right_w * 0.30
    dp_col_w    = (right_w - dp_bucket_w) / 4

    def draw_dp_table(times, top_y):
        yy = top_y
        col_xs = [rx, rx + dp_bucket_w, rx + dp_bucket_w + dp_col_w,
                  rx + dp_bucket_w + dp_col_w * 2, rx + dp_bucket_w + dp_col_w * 3]
        ax.add_patch(Rectangle((rx, yy), right_w, dp_header_h, facecolor="#1a6080",
                                edgecolor="none", zorder=2))
        ax.text(rx + 8, yy + dp_header_h / 2, "BUCKET", fontsize=7.5, fontweight="bold",
                color="white", ha="left", va="center", zorder=3)
        for i, t in enumerate(times):
            cx = col_xs[i + 1] + dp_col_w / 2
            ax.text(cx, yy + dp_header_h / 2, t, fontsize=7.5, fontweight="bold",
                    color="white", ha="center", va="center", zorder=3)
        yy += dp_header_h

        for bi, b in enumerate(DAY_PROCESS_BUCKETS):
            shaded = (bi % 2 == 0)
            row_bg = "#dce8f0" if shaded else "#ffffff"
            ax.add_patch(Rectangle((rx, yy), right_w, dp_row_h, facecolor=row_bg,
                                    edgecolor="none", zorder=2))
            ax.text(rx + 8, yy + dp_row_h / 2, b, fontsize=7.5, fontweight="bold",
                    color="#1a3a5c", ha="left", va="center", zorder=3)
            for i, t in enumerate(times):
                cx = col_xs[i + 1] + dp_col_w / 2
                ax.text(cx, yy + dp_row_h / 2, pct(history[b].get(t)), fontsize=7.5,
                        color="#333333", ha="center", va="center", zorder=3)
            ax.plot([rx, rx + right_w], [yy + dp_row_h, yy + dp_row_h], color="#eef0f3", linewidth=1, zorder=3)
            yy += dp_row_h

        ax.plot([rx, rx + right_w], [yy, yy], color="#1a3a5c", linewidth=1.5, zorder=3)
        ax.text(rx + 8, yy + dp_row_h / 2, "TOTAL", fontsize=7.5, fontweight="bold",
                color="#1a1a2e", ha="left", va="center", zorder=3)
        for i, t in enumerate(times):
            cx = col_xs[i + 1] + dp_col_w / 2
            ax.text(cx, yy + dp_row_h / 2, pct(day_process_total(history, t)), fontsize=7.5,
                    fontweight="bold", color="#1a1a2e", ha="center", va="center", zorder=3)
        yy += dp_row_h
        return yy

    ry = draw_dp_table(TIME_SLOTS[:4], ry)
    ry += dp_gap_between
    ry = draw_dp_table(TIME_SLOTS[4:], ry)

    buf = BytesIO()
    fig.savefig(buf, format="png", facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def build_excel_workbook(bucket_data, zone_data, history, timestamp_str, selected_time) -> bytes:
    """Builds the downloadable .xlsx. The 'DayProcess' sheet is what makes the
    time-slot history persist: re-uploading this exact file next time (as the
    'previously downloaded snapshot') lets earlier time-slots carry forward."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    arial          = Font(name="Arial")
    bold_font      = Font(name="Arial", bold=True)
    title_font     = Font(name="Arial", bold=True, size=14)
    header_font    = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill    = PatternFill("solid", fgColor="1A6080")
    pan_fill       = PatternFill("solid", fgColor="E6F4EA")
    zone_fill      = PatternFill("solid", fgColor="DCE8F0")
    center         = Alignment(horizontal="center")

    ws["A1"] = f"Collections Hourly Snapshot | Updates | {timestamp_str}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")
    ws["A2"] = f"Last Updated Slot: {selected_time}"
    ws["A2"].font = Font(name="Arial", italic=True, color="666666")

    # ── bucket overview
    r = 4
    for ci, label in enumerate(["Bucket", "Efficiency %", "Flow Value (Crs)"], start=1):
        c = ws.cell(r, ci, label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    r += 1
    for key, label, _, _ in BOX_CONFIGS:
        bd = bucket_data.get(key, {})
        ws.cell(r, 1, label).font = arial
        eff_cell = ws.cell(r, 2, bd.get('eff', 0) or 0)
        eff_cell.number_format = "0.00%"
        eff_cell.alignment = center
        flow_cell = ws.cell(r, 3, bd.get('flow', 0) or 0)
        flow_cell.number_format = "#,##0.00"
        flow_cell.alignment = center
        r += 1

    # ── sub-zone cut
    r += 1
    sz_header_row = r
    for ci, label in enumerate(["ZONE", "FRESH", "1-29", "30-59", "60-89"], start=1):
        c = ws.cell(sz_header_row, ci, label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    r += 1
    for zone_name, subs in ZONE_ORDER:
        display = zone_name.replace('_', ' ')
        is_pan  = zone_name == "PAN INDIA"
        fill    = pan_fill if is_pan else zone_fill
        ws.cell(r, 1, display).font = bold_font
        ws.cell(r, 1).fill = fill
        for ci, b in enumerate(['fresh', '1-29', '30-59', '60-89'], start=2):
            v = get_raw_val(zone_data, zone_name, b)
            cell = ws.cell(r, ci, v if v is not None else None)
            cell.fill = fill
            cell.alignment = center
            if v is not None:
                cell.number_format = "0.00%"
        r += 1
        for sub in subs:
            disp_sub = sub.replace('_', ' ')
            ws.cell(r, 1, disp_sub).font = arial
            for ci, b in enumerate(['fresh', '1-29', '30-59', '60-89'], start=2):
                v = get_raw_val(zone_data, sub, b)
                cell = ws.cell(r, ci, v if v is not None else None)
                cell.alignment = center
                if v is not None:
                    cell.number_format = "0.00%"
            r += 1

    ws.column_dimensions['A'].width = 22
    for col in "BCDE":
        ws.column_dimensions[col].width = 14

    # ── DayProcess sheet (history / persistence)
    dp = wb.create_sheet("DayProcess")
    dp.cell(1, 1, "Bucket").font = header_font
    dp.cell(1, 1).fill = header_fill
    dp.cell(1, 1).alignment = center
    for ci, t in enumerate(TIME_SLOTS, start=2):
        c = dp.cell(1, ci, t)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for ri, b in enumerate(DAY_PROCESS_BUCKETS, start=2):
        dp.cell(ri, 1, b).font = bold_font
        for ci, t in enumerate(TIME_SLOTS, start=2):
            v = history[b][t]
            cell = dp.cell(ri, ci, v)
            cell.alignment = center
            if v is not None:
                cell.number_format = "0.00%"

    total_row = 2 + len(DAY_PROCESS_BUCKETS)
    dp.cell(total_row, 1, "TOTAL").font = bold_font
    for ci, t in enumerate(TIME_SLOTS, start=2):
        tv = day_process_total(history, t)
        cell = dp.cell(total_row, ci, tv)
        cell.alignment = center
        if tv is not None:
            cell.number_format = "0.00%"

    dp.column_dimensions['A'].width = 10
    for col in "BCDEFGHI":
        dp.column_dimensions[col].width = 10

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── STREAMLIT UI ──────────────────────────────────────────────────────────────
_default_history = load_day_process_history(prev_file)
_default_idx = 0
for _i, _t in enumerate(TIME_SLOTS):
    if all(_default_history[_b].get(_t) is None for _b in DAY_PROCESS_BUCKETS):
        _default_idx = _i
        break
selected_time = st.selectbox("🕐 Select Time Slot", TIME_SLOTS, index=_default_idx)

if uploaded_file:
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        bucket_data, zone_data = parse_sheet3(wb)
    except Exception as e:
        st.error(f"Couldn't read the raw source Excel: {e}")
        st.stop()

    history = load_day_process_history(prev_file)
    history = update_day_process_history(history, bucket_data, selected_time)

    now = datetime.now()
    timestamp_str = (
        f"{now.day}-{now.strftime('%b-%y')} | "
        f"{now.strftime('%A')} | "
        f"{now.strftime('%I:%M %p').lstrip('0')}"
    )

    html = build_html(bucket_data, zone_data, timestamp_str, history)

    # ── live preview in streamlit
    st.components.v1.html(html, height=820, scrolling=True)

    st.markdown("---")
    st.markdown("### ⬇️ Download Snapshot")
    st.caption(
        f"Recording data into the **{selected_time}** slot. "
        "Download the Excel below and re-upload it next time (as file #2) to keep building up the day."
    )

    dl_col1, dl_col2 = st.columns(2)

    with dl_col1:
        if st.button("🖼️ Generate & Download Image"):
            with st.spinner("Rendering dashboard to high-res PNG..."):
                try:
                    png_bytes = render_dashboard_png(bucket_data, zone_data, timestamp_str, history)
                    st.success("Image ready!")
                    st.image(png_bytes, caption="Collections Hourly Snapshot", use_container_width=True)
                    st.download_button(
                        label="⬇️ Download PNG",
                        data=png_bytes,
                        file_name=f"collections_snapshot_{now.strftime('%Y%m%d_%H%M')}.png",
                        mime="image/png",
                    )
                except Exception as e:
                    st.error(f"Error generating image: {e}")

    with dl_col2:
        if st.button("📊 Generate & Download Excel"):
            with st.spinner("Building Excel workbook..."):
                try:
                    xlsx_bytes = build_excel_workbook(bucket_data, zone_data, history, timestamp_str, selected_time)
                    st.success("Excel ready!")
                    st.download_button(
                        label="⬇️ Download Excel",
                        data=xlsx_bytes,
                        file_name=f"collections_snapshot_{now.strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error(f"Error generating Excel: {e}")
else:
    st.info("📂 Please upload the raw source Excel file (.xlsx) to get started.")
