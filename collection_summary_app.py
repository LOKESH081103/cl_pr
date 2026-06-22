"""
Collection Dashboard Summary Generator
Upload the PAN INDIA .xlsb/.xlsx dashboard → Download formatted summary Excel
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import io
import re
from datetime import datetime

# ── Constants ──────────────────────────────────────────────────────────────
DARK_BLUE   = "FF1F3864"   # header row dark navy
MID_BLUE    = "FF2E5FA3"   # section sub-header
LIGHT_BLUE  = "FFD6E4F7"   # alt row
LIGHT_GRAY  = "FFF2F2F2"
WHITE       = "FFFFFFFF"
RED_TXT     = "FFCC0000"
GREEN_TXT   = "FF006600"
ORANGE_TXT  = "FFCC6600"
BORDER_COLOR = "FF8EA9C1"

THIN  = Side(style="thin",   color=BORDER_COLOR)
THICK = Side(style="medium", color="FF1F3864")

def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def thick_bottom():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THICK)

def hdr_font(size=10, bold=True, color="FFFFFFFF"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def cell_font(size=9, bold=False, color="FF000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)


# ── xlsb reader via pyxlsb (optional) or pandas ───────────────────────────
def read_proj_overall(file_bytes, filename: str) -> pd.DataFrame:
    """Read the PROJ-OVERALL sheet from xlsb or xlsx."""
    ext = filename.lower().split(".")[-1]
    try:
        if ext == "xlsb":
            try:
                import pyxlsb
                with pyxlsb.open_workbook(io.BytesIO(file_bytes)) as wb:
                    with wb.get_sheet("PROJ-OVERALL") as ws:
                        rows = [[c.v for c in r] for r in ws.rows()]
                df = pd.DataFrame(rows[1:], columns=rows[0])
            except ImportError:
                st.error("pyxlsb not installed. Run: pip install pyxlsb")
                st.stop()
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="PROJ-OVERALL", header=0)
    except Exception as e:
        st.error(f"Could not read PROJ-OVERALL sheet: {e}")
        st.stop()
    df.columns = [str(c).strip() if c is not None else f"Col_{i}"
                  for i, c in enumerate(df.columns)]
    return df


# ── Data extraction helpers ───────────────────────────────────────────────
def find_col(df: pd.DataFrame, patterns: list) -> str | None:
    """Find first column matching any pattern (case-insensitive)."""
    for pat in patterns:
        for c in df.columns:
            if re.search(pat, c, re.I):
                return c
    return None

def pct_fmt(val) -> str:
    try:
        v = float(val)
        return f"{v:.2f}%"
    except:
        return "---"

def num_fmt(val, dec=2) -> str:
    try:
        v = float(val)
        return f"{v:,.{dec}f}"
    except:
        return "---"

def arrow(today_val, yest_val):
    """Return ▲ or ▼ based on comparison."""
    try:
        return "▲" if float(today_val) >= float(yest_val) else "▼"
    except:
        return ""

def arrow_color(today_val, yest_val, higher_is_good=True):
    try:
        up = float(today_val) >= float(yest_val)
        good = up if higher_is_good else not up
        return GREEN_TXT if good else RED_TXT
    except:
        return "FF000000"


# ── Extract structured data from PROJ-OVERALL ────────────────────────────
def extract_data(df: pd.DataFrame) -> dict:
    """
    Pull Zone Efficiency, Norm/RB/RF, Bottom Sub-Zones from PROJ-OVERALL.
    Column names vary; we search by pattern.
    """
    col_today = find_col(df, [r"today", r"eff.*today", r"curr"])
    col_yest  = find_col(df, [r"yest", r"prev.*day", r"D-1"])
    col_lmtd  = find_col(df, [r"lmtd", r"jan.*26", r"month.*till"])
    col_aod   = find_col(df, [r"aod.*flow", r"flow.*value", r"aod.*cr"])
    col_proj  = find_col(df, [r"proj", r"projection"])
    col_bucket= find_col(df, [r"bucket", r"dpd", r"category", r"segment"])
    col_zone  = find_col(df, [r"zone", r"sub.*zone", r"region"])
    col_eff   = find_col(df, [r"eff.*%", r"efficiency"])

    return {
        "df": df,
        "cols": {
            "today": col_today, "yest": col_yest, "lmtd": col_lmtd,
            "aod": col_aod, "proj": col_proj, "bucket": col_bucket,
            "zone": col_zone, "eff": col_eff,
        }
    }


# ── Build output Excel ────────────────────────────────────────────────────
def write_cell(ws, row, col, value, font=None, fill=None, align=None,
               border=None, number_format=None, color=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    if color:  c.font = Font(name="Arial", size=9, color=color,
                              bold=c.font.bold if c.font else False)
    return c

def solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right_align():
    return Alignment(horizontal="right", vertical="center")


def generate_summary_excel(raw: dict, report_date: str) -> bytes:
    df   = raw["df"]
    cols = raw["cols"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = [22, 13, 13, 13, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ════════════════════════════════════════════════════════════════════════
    # HEADER BANNER
    # ════════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{row}:E{row}")
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=1,
                value=f"DAILY COLLECTIONS UPDATE | {report_date}   PAN INDIA")
    c.font  = Font(name="Arial", size=12, bold=True, color="FFFFFFFF")
    c.fill  = solid(DARK_BLUE)
    c.alignment = center_align()
    row += 1

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 1 – ZONE EFFICIENCY %
    # ════════════════════════════════════════════════════════════════════════
    def section_header(label, r):
        ws.merge_cells(f"A{r}:E{r}")
        ws.row_dimensions[r].height = 16
        c = ws.cell(row=r, column=1, value=label)
        c.font  = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        c.fill  = solid(MID_BLUE)
        c.alignment = left_align()
        c.border = thin_border()
        return r + 1

    def col_header_row(labels, r, bg=DARK_BLUE):
        ws.row_dimensions[r].height = 15
        for ci, lbl in enumerate(labels, 1):
            c = ws.cell(row=r, column=ci, value=lbl)
            c.font  = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
            c.fill  = solid(bg)
            c.alignment = center_align()
            c.border = thin_border()
        return r + 1

    def data_row(values, r, alt=False, arrow_cols=None, colors=None):
        ws.row_dimensions[r].height = 14
        bg = LIGHT_BLUE if alt else WHITE
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=r, column=ci, value=val)
            fnt = Font(name="Arial", size=9,
                       bold=(ci == 1),
                       color=colors[ci-1] if colors else "FF000000")
            c.font = fnt
            c.fill = solid(bg)
            c.alignment = center_align() if ci > 1 else left_align()
            c.border = thin_border()
        return r + 1

    row = section_header("1. Zone Efficiency%", row)
    row = col_header_row(["Bucket", "Today", "Yesterday", "LMTD", "AOD Flow Value\n(GA Crs.)"], row)

    # Try to pull real data; fall back to placeholder rows
    zone_eff_buckets = [
        ("Fresh",            cols["today"], cols["yest"], cols["lmtd"], cols["aod"]),
        ("1–29",             cols["today"], cols["yest"], cols["lmtd"], cols["aod"]),
        ("1–29 Norm%",       cols["today"], cols["yest"], cols["lmtd"], None),
        ("30-59",            cols["today"], cols["yest"], cols["lmtd"], cols["aod"]),
        ("60-89 (S3 CONCERN)",cols["today"],cols["yest"], cols["lmtd"], cols["aod"]),
    ]

    # Filter df rows for efficiency section
    bucket_col = cols["bucket"]
    today_col  = cols["today"]
    yest_col   = cols["yest"]
    lmtd_col   = cols["lmtd"]
    aod_col    = cols["aod"]

    BUCKET_PATTERNS = {
        "Fresh":              r"fresh",
        "1–29":               r"1.?29(?!.*norm)",
        "1–29 Norm%":         r"1.?29.*norm|norm.*1.?29",
        "30-59":              r"30.?59",
        "60-89 (S3 CONCERN)": r"60.?89|s3.*concern",
    }

    def lookup_row(df, bucket_col, pattern):
        if bucket_col is None:
            return None
        matches = df[df[bucket_col].astype(str).str.contains(pattern, case=False, na=False)]
        return matches.iloc[0] if len(matches) else None

    for i, (label, tc, yc, lc, ac) in enumerate(zone_eff_buckets):
        pat = BUCKET_PATTERNS.get(label, label)
        r_data = lookup_row(df, bucket_col, pat)

        today_v = r_data[tc]  if (r_data is not None and tc) else "---"
        yest_v  = r_data[yc]  if (r_data is not None and yc) else "---"
        lmtd_v  = r_data[lc]  if (r_data is not None and lc) else "---"
        aod_v   = r_data[ac]  if (r_data is not None and ac) else "---"

        arr = arrow(today_v, yest_v)
        a_color = arrow_color(today_v, yest_v) if label != "1–29 Norm%" else GREEN_TXT

        today_str = f"{arr} {pct_fmt(today_v)}" if label != "AOD" else pct_fmt(today_v)
        yest_str  = pct_fmt(yest_v)
        lmtd_str  = pct_fmt(lmtd_v)
        aod_str   = num_fmt(aod_v) if ac else "---"

        colors = ["FF000000", a_color, "FF000000", "FF000000", "FF000000"]
        row = data_row([label, today_str, yest_str, lmtd_str, aod_str],
                       row, alt=(i % 2 == 1), colors=colors)

    row += 1   # spacer

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 2 – NORM / RB / RF METRICS
    # ════════════════════════════════════════════════════════════════════════
    row = section_header("2. Norm / RB / RF Metrics", row)
    row = col_header_row(["Bucket", "Today", "Yesterday", "LMTD", "Projection Crs."], row)

    METRIC_PATTERNS = {
        "1-29 Norm":           r"1.?29.*norm|norm.*1.?29",
        "Stage 2 Roll Back":   r"stage.?2.*roll|roll.*back.*2",
        "Stage 3 Concern Flow":r"stage.?3.*concern.*flow|s3.*concern.*flow",
        "Stage 3 Roll Back":   r"stage.?3.*roll|roll.*back.*3",
    }

    proj_col = cols["proj"]

    for i, (label, pat) in enumerate(METRIC_PATTERNS.items()):
        r_data = lookup_row(df, bucket_col, pat)

        today_v = r_data[today_col] if (r_data is not None and today_col) else "---"
        yest_v  = r_data[yest_col]  if (r_data is not None and yest_col)  else "---"
        lmtd_v  = r_data[lmtd_col]  if (r_data is not None and lmtd_col)  else "---"
        proj_v  = r_data[proj_col]  if (r_data is not None and proj_col)  else "---"

        arr     = arrow(today_v, yest_v)
        a_color = arrow_color(today_v, yest_v)

        colors = ["FF000000", a_color, "FF000000", "FF000000", "FF000000"]
        row = data_row(
            [label, f"{arr} {num_fmt(today_v)}", num_fmt(yest_v),
             num_fmt(lmtd_v), num_fmt(proj_v)],
            row, alt=(i % 2 == 1), colors=colors
        )

    # footnotes
    ws.row_dimensions[row].height = 12
    ws.merge_cells(f"A{row}:E{row}")
    fn = ws.cell(row=row, column=1,
                 value="(a) * Agreement level S3 concern,  (b) Efficiency% excluding hold cases.")
    fn.font = Font(name="Arial", size=8, italic=True, color="FF555555")
    fn.alignment = left_align()
    row += 2

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 3 – BOTTOM SUB-ZONES | EFFICIENCY %
    # ════════════════════════════════════════════════════════════════════════
    row = section_header("Bottom Sub-Zones | Efficiency%", row)
    row = col_header_row(["Bucket", "Bottom 1", "", "Bottom 2", ""], row)

    BOTTOM_EFF_BUCKETS = [
        ("Fresh",           BUCKET_PATTERNS["Fresh"]),
        ("1-29 (S2 Concern)",r"1.?29"),
        ("1-29 Norm",       BUCKET_PATTERNS["1–29 Norm%"]),
        ("30-59",           BUCKET_PATTERNS["30-59"]),
        ("60-89 (S3 Concern)", BUCKET_PATTERNS["60-89 (S3 CONCERN)"]),
    ]

    zone_col = cols["zone"]
    eff_col  = cols["eff"]

    def bottom_zones(df, bucket_col, zone_col, eff_col, bucket_pat, n=2):
        """Return bottom N (zone, eff%) pairs for a bucket."""
        if bucket_col is None or zone_col is None or eff_col is None:
            return [("---", "---")] * n
        sub = df[df[bucket_col].astype(str).str.contains(bucket_pat, case=False, na=False)].copy()
        sub[eff_col] = pd.to_numeric(sub[eff_col], errors="coerce")
        sub = sub.dropna(subset=[eff_col]).sort_values(eff_col)
        results = []
        for _, r in sub.head(n).iterrows():
            results.append((str(r[zone_col]), pct_fmt(r[eff_col])))
        while len(results) < n:
            results.append(("---", "---"))
        return results

    for i, (label, pat) in enumerate(BOTTOM_EFF_BUCKETS):
        bots = bottom_zones(df, bucket_col, zone_col, eff_col, pat, 2)
        b1_zone, b1_eff = bots[0]
        b2_zone, b2_eff = bots[1]
        row = data_row(
            [label, f"{b1_zone} {b1_eff}", "", f"{b2_zone} {b2_eff}", ""],
            row, alt=(i % 2 == 1)
        )

    row += 1

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 4 – BOTTOM SUB-ZONES | AOD FLOW VALUE
    # ════════════════════════════════════════════════════════════════════════
    row = section_header("Bottom Sub-Zones | AOD Flow Value Crs.", row)
    row = col_header_row(["Bucket", "Bottom 1", "", "Bottom 2", ""], row)

    def bottom_zones_aod(df, bucket_col, zone_col, aod_col, bucket_pat, n=2):
        if bucket_col is None or zone_col is None or aod_col is None:
            return [("---", "---")] * n
        sub = df[df[bucket_col].astype(str).str.contains(bucket_pat, case=False, na=False)].copy()
        sub[aod_col] = pd.to_numeric(sub[aod_col], errors="coerce")
        sub = sub.dropna(subset=[aod_col]).sort_values(aod_col)
        results = []
        for _, r in sub.head(n).iterrows():
            results.append((str(r[zone_col]), f"₹ {num_fmt(r[aod_col])}"))
        while len(results) < n:
            results.append(("---", "---"))
        return results

    BOTTOM_AOD_BUCKETS = [
        ("Fresh",              BUCKET_PATTERNS["Fresh"]),
        ("1-29 (S2 Concern)",  r"1.?29"),
        ("30-59",              BUCKET_PATTERNS["30-59"]),
        ("60-89 (S3 Concern)", BUCKET_PATTERNS["60-89 (S3 CONCERN)"]),
    ]

    for i, (label, pat) in enumerate(BOTTOM_AOD_BUCKETS):
        bots = bottom_zones_aod(df, bucket_col, zone_col, aod_col, pat, 2)
        b1_zone, b1_val = bots[0]
        b2_zone, b2_val = bots[1]
        row = data_row(
            [label, f"{b1_zone} {b1_val}", "", f"{b2_zone} {b2_val}", ""],
            row, alt=(i % 2 == 1)
        )

    # ── Freeze panes & print setup ─────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.print_title_rows = "1:2"

    # ── Save to bytes ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Collection Dashboard Summary",
    page_icon="📊",
    layout="centered"
)

st.title("📊 Daily Collections Summary Generator")
st.markdown(
    "Upload the **PAN INDIA Collection Dashboard** (`.xlsb` or `.xlsx`) "
    "and download the formatted summary — matching the screenshot layout."
)

with st.sidebar:
    st.header("⚙️ Settings")
    report_date = st.date_input("Report Date", value=datetime.today())
    date_str = report_date.strftime("%-d-%b-%y")   # e.g. 27-May-26

uploaded = st.file_uploader(
    "Upload Dashboard (.xlsb or .xlsx)",
    type=["xlsb", "xlsx"],
    help="Must contain a sheet named PROJ-OVERALL"
)

if uploaded:
    with st.spinner("Reading PROJ-OVERALL sheet…"):
        file_bytes = uploaded.read()
        raw_df = read_proj_overall(file_bytes, uploaded.name)

    st.success(f"✅ Loaded **{len(raw_df):,}** rows from PROJ-OVERALL")

    # Column mapping UI
    with st.expander("🔎 Column Mapping (auto-detected — adjust if needed)"):
        all_cols = ["(none)"] + list(raw_df.columns)
        def col_pick(label, patterns):
            default = find_col(raw_df, patterns) or "(none)"
            idx = all_cols.index(default) if default in all_cols else 0
            return st.selectbox(label, all_cols, index=idx)

        today_c  = col_pick("Today / Efficiency% column",    [r"today", r"curr"])
        yest_c   = col_pick("Yesterday column",               [r"yest", r"D-1"])
        lmtd_c   = col_pick("LMTD column",                    [r"lmtd", r"month"])
        aod_c    = col_pick("AOD Flow Value column",           [r"aod", r"flow.*value"])
        proj_c   = col_pick("Projection Crs. column",          [r"proj"])
        bucket_c = col_pick("Bucket / Category column",        [r"bucket", r"dpd", r"segment"])
        zone_c   = col_pick("Sub-Zone / Zone column",          [r"zone", r"sub.*zone"])
        eff_c    = col_pick("Efficiency % (for bottom zones)", [r"eff.*%", r"efficiency"])

    # Override auto-detect with user picks
    def none_to_none(s):
        return None if s == "(none)" else s

    cols_override = {
        "today": none_to_none(today_c), "yest": none_to_none(yest_c),
        "lmtd":  none_to_none(lmtd_c),  "aod":  none_to_none(aod_c),
        "proj":  none_to_none(proj_c),   "bucket": none_to_none(bucket_c),
        "zone":  none_to_none(zone_c),   "eff":  none_to_none(eff_c),
    }

    raw = {"df": raw_df, "cols": cols_override}

    # Preview
    with st.expander("📋 Preview raw data (first 20 rows)"):
        st.dataframe(raw_df.head(20))

    if st.button("🚀 Generate Summary Excel", type="primary"):
        with st.spinner("Building formatted Excel…"):
            xlsx_bytes = generate_summary_excel(raw, date_str)

        out_name = f"Collections_Summary_{report_date.strftime('%d%b%Y')}.xlsx"
        st.download_button(
            label="⬇️ Download Summary Excel",
            data=xlsx_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.balloons()
else:
    st.info("👆 Upload a dashboard file to get started.")

st.markdown("---")
st.caption("Built for PAN INDIA Daily Collections | PROJ-OVERALL sheet")
