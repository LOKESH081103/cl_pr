"""
Collection Summary Generator
Input:  PROJ-OVERALL tab saved as TSV or Excel
Output: Formatted summary Excel matching the screenshot layout
"""

import sys
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── COLUMN INDEX MAP (0-based) ───────────────────────────────────────────────
# Col 0  = Zone / Sub-Zone name
# Col 1  = AOD Total
# Col 2  = Collected Total
C = {
    'zone': 0, 'aod_total': 1, 'coll_total': 2,
    # Fresh
    'fresh_today': 3,  'fresh_arrow': 4,  'fresh_delta': 5,
    'fresh_lmtd':  6,  'fresh_norm':  7,  'fresh_aod':   8,
    # 1-29
    '129_today':  11, '129_arrow': 12, '129_delta': 13,
    '129_lmtd':   14, '129_aod':   15, '129_norm':  16,
    # 1-29 Norm%
    '129n_today': 19, '129n_aod': 20,
    '129n_lmtd':  24,
    # 30-59
    '3059_aod':   28, '3059_today': 31, '3059_arrow': 32,
    '3059_delta': 33, '3059_lmtd':  34, '3059_norm':  35,
    # 60-89
    '6089_aod':   39, '6089_today': 42, '6089_arrow': 43,
    '6089_delta': 44, '6089_lmtd':  45, '6089_norm':  47,
}

# ─── STYLE HELPERS ────────────────────────────────────────────────────────────
DARK_NAVY  = "1F3864"
MID_BLUE   = "2E4D8E"
STEEL_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F7"
ALT_WHITE  = "FFFFFF"
ALT_GRAY   = "F2F7FC"
RED        = "C00000"
GREEN      = "375623"
ARROW_RED  = "FF0000"
ARROW_GRN  = "00B050"
BORDER_CLR = "8EA9C1"

def solid(hex6):
    return PatternFill("solid", fgColor=hex6)

def bdr(thin=True):
    s = Side(style="thin" if thin else "medium", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(sz=10, bold=True):
    return Font(name="Arial", size=sz, bold=bold, color="FFFFFF")

def dat_font(sz=9, bold=False, color="000000"):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def ctr(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def lft(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def rgt():
    return Alignment(horizontal="right", vertical="center")

# ─── DATA HELPERS ─────────────────────────────────────────────────────────────
def v(row, col_key):
    """Get value from row by column key."""
    idx = C[col_key]
    if idx >= len(row):
        return None
    val = row[idx]
    if pd.isna(val) if not isinstance(val, str) else (val.strip() == ''):
        return None
    return val

def pct(val):
    if val is None: return "---"
    s = str(val).strip().rstrip('%')
    try:
        f = float(s)
        return f"{f:.2f}%"
    except:
        return str(val)

def amt(val, dec=2):
    if val is None: return "---"
    s = str(val).strip().replace(',','')
    try:
        f = float(s)
        return f"{f:,.2f}"
    except:
        return str(val)

def arrow_str(val):
    if val is None: return ""
    s = str(val).strip()
    if s in ('↑','▲'): return "▲"
    if s in ('↓','▼'): return "▼"
    # derive from delta
    try:
        f = float(s.replace('%',''))
        return "▲" if f >= 0 else "▼"
    except:
        return s

def arrow_color(arrow):
    if arrow == "▲": return ARROW_GRN
    if arrow == "▼": return ARROW_RED
    return "000000"

def delta_str(val):
    if val is None: return ""
    s = str(val).strip()
    try:
        f = float(s.replace('%',''))
        sign = "+" if f >= 0 else ""
        return f"{sign}{f:.2f}%"
    except:
        return s

# ─── EXCEL WRITER HELPERS ──────────────────────────────────────────────────────
def wc(ws, r, c, val, font=None, fill=None, align=None, border=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    return cell

def section_hdr(ws, r, label):
    ws.merge_cells(f"A{r}:H{r}")
    ws.row_dimensions[r].height = 17
    cell = ws.cell(row=r, column=1, value=label)
    cell.font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    cell.fill  = solid(DARK_NAVY)
    cell.alignment = lft()
    cell.border = bdr()
    return r + 1

def col_hdrs(ws, r, labels, bg=MID_BLUE, merge_pairs=None):
    ws.row_dimensions[r].height = 15
    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(row=r, column=ci, value=lbl)
        c.font  = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill  = solid(bg)
        c.alignment = ctr(wrap=True)
        c.border = bdr()
    return r + 1

def data_row_s1(ws, r, bucket, today, lmtd, delta, arrow, aod, alt):
    """Section 1 data row: Bucket | Today (arrow delta) | LMTD | AOD"""
    bg = ALT_GRAY if alt else ALT_WHITE
    ws.row_dimensions[r].height = 14
    arw = arrow_str(arrow or delta)
    aclr = arrow_color(arw)
    dlt  = delta_str(delta)

    # Col A: bucket label
    c = ws.cell(row=r, column=1, value=bucket)
    c.font = dat_font(bold=True); c.fill = solid(bg); c.alignment = lft(); c.border = bdr()

    # Col B: today%
    c = ws.cell(row=r, column=2, value=pct(today))
    c.font = dat_font(); c.fill = solid(bg); c.alignment = ctr(); c.border = bdr()

    # Col C: arrow
    c = ws.cell(row=r, column=3, value=arw)
    c.font = Font(name="Arial", size=9, color=aclr, bold=True)
    c.fill = solid(bg); c.alignment = ctr(); c.border = bdr()

    # Col D: delta
    c = ws.cell(row=r, column=4, value=dlt)
    c.font = Font(name="Arial", size=9, color=aclr)
    c.fill = solid(bg); c.alignment = ctr(); c.border = bdr()

    # Col E: LMTD
    c = ws.cell(row=r, column=5, value=pct(lmtd))
    c.font = dat_font(); c.fill = solid(bg); c.alignment = ctr(); c.border = bdr()

    # Col F-H: AOD Flow Value (merged)
    ws.merge_cells(f"F{r}:H{r}")
    c = ws.cell(row=r, column=6, value=amt(aod))
    c.font = dat_font(); c.fill = solid(bg); c.alignment = rgt(); c.border = bdr()

    return r + 1

def data_row_bottom(ws, r, bucket, b1_zone, b1_val, b2_zone, b2_val, alt, is_aod=False):
    """Bottom sub-zones row."""
    bg = ALT_GRAY if alt else ALT_WHITE
    ws.row_dimensions[r].height = 14
    prefix = "₹ " if is_aod else ""

    # bucket
    c = ws.cell(row=r, column=1, value=bucket)
    c.font = dat_font(bold=True); c.fill = solid(bg); c.alignment = lft(); c.border = bdr()

    # bottom 1 zone + value (cols B-D merged)
    ws.merge_cells(f"B{r}:D{r}")
    b1_text = f"{b1_zone}  {prefix}{b1_val}" if b1_zone != "---" else "---"
    c = ws.cell(row=r, column=2, value=b1_text)
    c.font = dat_font(); c.fill = solid(bg); c.alignment = lft(); c.border = bdr()

    # bottom 2 zone + value (cols E-H merged)
    ws.merge_cells(f"E{r}:H{r}")
    b2_text = f"{b2_zone}  {prefix}{b2_val}" if b2_zone != "---" else "---"
    c = ws.cell(row=r, column=5, value=b2_text)
    c.font = dat_font(); c.fill = solid(bg); c.alignment = lft(); c.border = bdr()

    return r + 1

# ─── BOTTOM ZONE FINDER ───────────────────────────────────────────────────────
def get_bottom2(subzone_rows, val_col_key, fmt_fn, exclude_zone_patterns=None):
    """Find 2 sub-zones with lowest values for a given column."""
    pairs = []
    for row in subzone_rows:
        zone = str(row[C['zone']]).strip()
        val = row[C[val_col_key]]
        try:
            fval = float(str(val).replace('%','').replace(',','').strip())
            pairs.append((zone, fval))
        except:
            pass
    pairs.sort(key=lambda x: x[1])
    result = []
    for zone, fval in pairs[:2]:
        result.append((zone, fmt_fn(fval)))
    while len(result) < 2:
        result.append(("---", "---"))
    return result

# ─── MAIN BUILD FUNCTION ──────────────────────────────────────────────────────
def build_excel(df: pd.DataFrame, report_date: str) -> bytes:
    # Identify PAN INDIA row and sub-zone rows (exclude zone totals like NORTH, EAST etc.)
    # Zone name patterns: zone totals = all caps no underscore; sub-zones = contain _
    all_rows = df.values.tolist()

    pan_row = None
    zone_rows = {}    # e.g. {'NORTH': row, ...}
    subzone_rows = {} # e.g. {'NORTH_1': row, ...}

    for row in all_rows:
        name = str(row[0]).strip().upper()
        if name == "PAN INDIA":
            pan_row = row
        elif "_" in name:
            subzone_rows[name] = row
        elif name and name not in ("ZONE", "SUB ZONE", ""):
            zone_rows[name] = row

    if pan_row is None:
        # fallback: last row
        pan_row = all_rows[-1]

    sub_list = list(subzone_rows.values())  # for bottom zone calcs

    # ── Create workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Summary"
    ws.sheet_view.showGridLines = False

    # Column widths: A=22, B=10, C=5, D=10, E=10, F=8, G=8, H=8
    widths = [22, 10, 5, 10, 10, 9, 9, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ════════════════════════════════════════════════════════════════════════
    # TITLE BANNER
    # ════════════════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{row}:H{row}")
    ws.row_dimensions[row].height = 24
    c = ws.cell(row=row, column=1,
                value=f"DAILY COLLECTIONS UPDATE | {report_date}          PAN INDIA")
    c.font  = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill  = solid(DARK_NAVY)
    c.alignment = ctr()
    c.border = bdr()
    row += 1

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 1 – ZONE EFFICIENCY %
    # ════════════════════════════════════════════════════════════════════════
    row = section_hdr(ws, row, "1. Zone Efficiency%")
    row = col_hdrs(ws, row,
                   ["Bucket", "Today", "▲▼", "vs LMTD", "LMTD\n(Jan-26)", "AOD Flow", "Value", "(GA Crs.)"])

    buckets_s1 = [
        ("Fresh",             'fresh_today', 'fresh_lmtd', 'fresh_delta', 'fresh_arrow', 'fresh_aod'),
        ("1–29",              '129_today',   '129_lmtd',   '129_delta',   '129_arrow',   '129_aod'),
        ("1–29 Norm%",        '129n_today',  '129n_lmtd',  None,          None,          None),
        ("30–59",             '3059_today',  '3059_lmtd',  '3059_delta',  '3059_arrow',  '3059_aod'),
        ("60–89 (S3 Concern)","6089_today",  '6089_lmtd',  '6089_delta',  '6089_arrow',  '6089_aod'),
    ]

    for i, (label, tc, lc, dc, ac, aodc) in enumerate(buckets_s1):
        row = data_row_s1(
            ws, row, label,
            today = v(pan_row, tc),
            lmtd  = v(pan_row, lc) if lc else None,
            delta = v(pan_row, dc) if dc else None,
            arrow = v(pan_row, ac) if ac else None,
            aod   = v(pan_row, aodc) if aodc else None,
            alt   = (i % 2 == 1)
        )

    row += 1  # spacer

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 2 – ZONE NORM %  (Norm% per bucket across zones)
    # ════════════════════════════════════════════════════════════════════════
    row = section_hdr(ws, row, "2. Zone Norm %")
    row = col_hdrs(ws, row, ["Bucket", "PAN INDIA\nNorm%", "NORTH\nNorm%", "EAST\nNorm%",
                              "SOUTH\nNorm%", "WEST\nNorm%", "", ""])

    norm_buckets = [
        ("Fresh",        'fresh_norm'),
        ("1–29",         '129_norm'),
        ("30–59",        '3059_norm'),
        ("60–89 (S3 C)", '6089_norm'),
    ]

    zone_order = ["NORTH","EAST","SOUTH","WEST"]
    for i, (label, nc) in enumerate(norm_buckets):
        bg = ALT_GRAY if i % 2 == 1 else ALT_WHITE
        ws.row_dimensions[row].height = 14

        c = ws.cell(row=row, column=1, value=label)
        c.font = dat_font(bold=True); c.fill = solid(bg); c.alignment = lft(); c.border = bdr()

        c = ws.cell(row=row, column=2, value=pct(v(pan_row, nc)))
        c.font = dat_font(); c.fill = solid(bg); c.alignment = ctr(); c.border = bdr()

        for zi, zname in enumerate(zone_order, 3):
            zrow = zone_rows.get(zname)
            val = pct(v(zrow, nc)) if zrow else "---"
            c = ws.cell(row=row, column=zi, value=val)
            c.font = dat_font(); c.fill = solid(bg); c.alignment = ctr(); c.border = bdr()

        # empty cols 7-8
        for col in [7, 8]:
            c = ws.cell(row=row, column=col, value="")
            c.fill = solid(bg); c.border = bdr()

        row += 1

    row += 1

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 3 – BOTTOM SUB-ZONES | EFFICIENCY %
    # ════════════════════════════════════════════════════════════════════════
    row = section_hdr(ws, row, "Bottom Sub-Zones | Efficiency%")

    # Header row
    ws.row_dimensions[row].height = 15
    ws.merge_cells(f"A{row}:A{row}")
    c = ws.cell(row=row, column=1, value="Bucket")
    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    c.fill = solid(MID_BLUE); c.alignment = ctr(); c.border = bdr()

    ws.merge_cells(f"B{row}:D{row}")
    c = ws.cell(row=row, column=2, value="Bottom 1")
    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    c.fill = solid(MID_BLUE); c.alignment = ctr(); c.border = bdr()

    ws.merge_cells(f"E{row}:H{row}")
    c = ws.cell(row=row, column=5, value="Bottom 2")
    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    c.fill = solid(MID_BLUE); c.alignment = ctr(); c.border = bdr()
    row += 1

    eff_buckets = [
        ("Fresh",             'fresh_today'),
        ("1–29 (S2 Concern)", '129_today'),
        ("1–29 Norm%",        '129n_today'),
        ("30–59",             '3059_today'),
        ("60–89 (S3 Concern)","6089_today"),
    ]

    for i, (label, eff_col) in enumerate(eff_buckets):
        b2 = get_bottom2(sub_list, eff_col, lambda f: f"{f:.2f}%")
        row = data_row_bottom(ws, row, label,
                              b2[0][0], b2[0][1], b2[1][0], b2[1][1], alt=(i%2==1))

    row += 1

    # ════════════════════════════════════════════════════════════════════════
    # SECTION 4 – BOTTOM SUB-ZONES | AOD FLOW VALUE
    # ════════════════════════════════════════════════════════════════════════
    row = section_hdr(ws, row, "Bottom Sub-Zones | AOD Flow Value Crs.")

    ws.row_dimensions[row].height = 15
    c = ws.cell(row=row, column=1, value="Bucket")
    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    c.fill = solid(MID_BLUE); c.alignment = ctr(); c.border = bdr()
    ws.merge_cells(f"B{row}:D{row}")
    c = ws.cell(row=row, column=2, value="Bottom 1")
    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    c.fill = solid(MID_BLUE); c.alignment = ctr(); c.border = bdr()
    ws.merge_cells(f"E{row}:H{row}")
    c = ws.cell(row=row, column=5, value="Bottom 2")
    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    c.fill = solid(MID_BLUE); c.alignment = ctr(); c.border = bdr()
    row += 1

    aod_buckets = [
        ("Fresh",             'fresh_aod'),
        ("1–29 (S2 Concern)", '129_aod'),
        ("30–59",             '3059_aod'),
        ("60–89 (S3 Concern)","6089_aod"),
    ]

    for i, (label, aod_col) in enumerate(aod_buckets):
        b2 = get_bottom2(sub_list, aod_col, lambda f: f"{f:,.2f}")
        row = data_row_bottom(ws, row, label,
                              b2[0][0], b2[0][1], b2[1][0], b2[1][1],
                              alt=(i%2==1), is_aod=True)

    # ── Print setup ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── CLI ENTRY POINT ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse, os
    from datetime import datetime

    parser = argparse.ArgumentParser(description="Generate collection summary Excel")
    parser.add_argument("input",  help="PROJ-OVERALL file (.xlsx/.xls/.tsv/.csv)")
    parser.add_argument("--date", default=datetime.today().strftime("%-d-%b-%y"),
                        help="Report date string, e.g. '22-Jun-26'")
    parser.add_argument("--sheet", default="PROJ-OVERALL",
                        help="Sheet name if Excel input")
    parser.add_argument("--output", default="Collections_Summary.xlsx")
    args = parser.parse_args()

    ext = args.input.lower().split(".")[-1]
    if ext in ("xlsx","xls","xlsb"):
        df = pd.read_excel(args.input, sheet_name=args.sheet, header=None)
    elif ext == "tsv":
        df = pd.read_csv(args.input, sep="\t", header=None)
    else:
        df = pd.read_csv(args.input, header=None)

    print(f"Loaded {len(df)} rows, {len(df.columns)} cols from {args.input}")

    xlsx = build_excel(df, args.date)
    with open(args.output, "wb") as f:
        f.write(xlsx)
    print(f"✅ Written: {args.output}")
